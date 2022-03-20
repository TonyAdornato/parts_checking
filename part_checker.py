"""
This program takes the Durr part number and enters it in the stores web page
if stores returns a value it will paste that into the column to the right
It will then save the file under a new name

This program has a hard coded path for the directory

Ford parts all have the form K19-#######
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import os
import re


def program_start():
  
    # Set up the Excel logic
    os.chdir('C:\\Users\\AADORNAT\\Desktop') # Set directory for workbook
    wb = openpyxl.load_workbook('Ford KTP Sealer Robots Spare Parts.xlsx')
    sheet = wb[wb.sheetnames[0]]    #calls the sheet at index 0

    # Set up the Web page
    storesURL = 'https://webapps.ktp.ford.com/sfrquery/SFRquery.asp'
    browser = webdriver.Chrome()  # Chromedriver in Path
    browser.implicitly_wait(10) # set the program to try the web page for 10 seconds
    browser.get(storesURL)      # open the stores page

    active_button = browser.find_element(By.CSS_SELECTOR, '#bySelection > div:nth-child(4) > div > span')   # Entrance Screen
    active_button.click() 

    # Return the excel workbook and the webbrowser
    return wb, browser


# Check Part. Accepts a part number and returns a list of results
def check_part(part_number, browser):

    # Find search bar and type part number into it 
    search_elem = browser.find_element(By.CSS_SELECTOR, 
        'body > form > table:nth-child(2) > tbody > tr:nth-child(2) > td:nth-child(2) > input:nth-child(2)')
    search_elem.click()
    search_elem.send_keys(part_number)

    # Find search button and click it
    button = browser.find_element(By.CSS_SELECTOR, 
        'body > form > table:nth-child(2) > tbody > tr:nth-child(3) > td > input[type=button]:nth-child(1)')
    button.click()  # Click search button

    stores_table = browser.find_elements(By.CSS_SELECTOR, "#GenStoreTable > tbody") # create a list of the search results
    search_elem.clear()             # clear the entry from the search bar
    return stores_table             # return the table


def write_to_excel(wb, i, stores_list):
    
    sheet = wb[wb.sheetnames[0]]
    part_num_regex = re.compile(r'K19-\d+')         # set a regex to find stores number from input

    for x in stores_list:                           #loop over the data in the stores list
        mo = part_num_regex.findall(x.text)              # break the raw output into usable chunks
        if mo != None:
            sheet.cell(row=i, column=15).value = ", ".join(mo)      # write stores number
            sheet.cell(row=i, column=16).value = x.text             # write raw output
            print(mo)                                               # print so user can see it


def look_for_parts(wb, browser):
    
    sheet = wb[wb.sheetnames[0]]                #calls the first sheet (at index 0)
    data_length = len(sheet['G'])               # get number of lines to search
    
    for i in range(1, data_length+1):           # loop over all rows in input sheet
        c = sheet.cell(row=i, column=7).value   # get Durr part number from sheet
        if c != None:
            print(c)                            # for display purposes. user can see progress of the search
            stores_list = check_part(c, browser)
            write_to_excel(wb, i, stores_list)


# Main program
if __name__ == "__main__":
    try:
        wb, browser = program_start()
        look_for_parts(wb, browser)
    finally:
        wb.save('output.xlsx')
